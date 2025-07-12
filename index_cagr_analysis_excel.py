#!/usr/bin/env python3
"""
CAGR Analysis Excel Generator
Creates an Excel sheet with CAGR calculations for all NIFTY indices
Rows: Index names 
Columns: Number of years (1-20)
Cell values: CAGR percentage or "NA" for unavailable data
"""

import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')

from yahoo_nifty_data import get_available_indices

def calculate_cagr(start_price, end_price, years):
    """Calculate CAGR given start price, end price, and number of years"""
    if start_price <= 0 or end_price <= 0 or years <= 0:
        return None
    
    try:
        cagr = (end_price / start_price) ** (1 / years) - 1
        return cagr * 100  # Convert to percentage
    except:
        return None

def get_index_data_for_period(symbol, years_back):
    """Get index data for a specific period"""
    try:
        # Calculate start and end dates
        end_date = datetime.now()
        start_date = end_date - timedelta(days=int(years_back * 365.25) + 1)  # Add a buffer day
        
        # Get data
        ticker = yf.Ticker(symbol)
        hist = ticker.history(start=start_date, end=end_date)
        
        if hist.empty or len(hist) < 2:
            return None, None, None
        
        # Get the earliest and latest prices
        start_price = hist['Close'].iloc[0]
        end_price = hist['Close'].iloc[-1]
        
        # Calculate actual years between first and last data points
        actual_years = (hist.index[-1] - hist.index[0]).days / 365.25
        
        return start_price, end_price, actual_years
    
    except Exception as e:
        print(f"Error getting data for {symbol}: {str(e)[:50]}")
        return None, None, None

def test_index_availability(symbol):
    """Test if an index is available on Yahoo Finance"""
    try:
        ticker = yf.Ticker(symbol)
        hist = ticker.history(period="5d")
        return not hist.empty and len(hist) > 0
    except:
        return False

def main():
    print("=" * 80)
    print("CAGR ANALYSIS EXCEL GENERATOR")
    print("=" * 80)
    
    # Get all available indices dynamically
    indices = get_available_indices()
    
    # Time periods (1 to 20 years) - these will be columns
    years_list = list(range(1, 21))
    
    # Initialize results DataFrame - TRANSPOSED structure
    # Rows = Index names, Columns = Years
    results_df = pd.DataFrame(index=list(indices.keys()), columns=[f"{y}Y" for y in years_list])
    results_df.index.name = 'Index'
    
    print(f"Analyzing {len(indices)} indices for {len(years_list)} time periods...")
    print(f"Total calculations: {len(indices) * len(years_list)}")
    print("-" * 80)
    
    # Track availability
    available_indices = {}
    unavailable_indices = {}
    
    # Process each index
    for i, (index_name, symbol) in enumerate(indices.items(), 1):
        print(f"[{i:2d}/{len(indices)}] Processing: {index_name}")
        print(f"                    Symbol: {symbol}")
        
        # Test availability first
        is_available = test_index_availability(symbol)
        
        if not is_available:
            print(f"    ❌ Index not available on Yahoo Finance")
            unavailable_indices[index_name] = symbol
            # Fill all years with "NA"
            for year in years_list:
                results_df.loc[index_name, f"{year}Y"] = "NA"
            print()
            continue
        
        print(f"    ✅ Index available - calculating CAGR...")
        available_indices[index_name] = symbol
        
        # Calculate CAGR for each time period
        for years in years_list:
            start_price, end_price, actual_years = get_index_data_for_period(symbol, years)
            
            if start_price is not None and end_price is not None and actual_years is not None:
                # Use actual years for more accurate CAGR calculation
                cagr = calculate_cagr(start_price, end_price, actual_years)
                if cagr is not None:
                    results_df.loc[index_name, f"{years}Y"] = round(cagr, 2)
                    print(f"      {years}Y: {cagr:6.2f}%")
                else:
                    results_df.loc[index_name, f"{years}Y"] = "NA"
                    print(f"      {years}Y: NA (calc error)")
            else:
                results_df.loc[index_name, f"{years}Y"] = "NA"
                print(f"      {years}Y: NA (no data)")
        
        print()
    
    # Print availability summary
    print("=" * 80)
    print("INDEX AVAILABILITY SUMMARY")
    print("=" * 80)
    print(f"✅ Available indices: {len(available_indices)}")
    print(f"❌ Unavailable indices: {len(unavailable_indices)}")
    print(f"📊 Total data availability: {len(available_indices)}/{len(indices)} ({len(available_indices)/len(indices)*100:.1f}%)")
    
    if unavailable_indices:
        print(f"\n❌ UNAVAILABLE INDICES ({len(unavailable_indices)}):")
        for name, symbol in unavailable_indices.items():
            print(f"  • {name} ({symbol})")
    
    # Create Excel file with formatting
    print("\nCreating Excel file...")
    filename = f"NIFTY_CAGR_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Write the main data
        results_df.to_excel(writer, sheet_name='CAGR Analysis', index=True)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['CAGR Analysis']
        
        # Apply formatting
        apply_excel_formatting(worksheet, results_df)
        
        # Create summary sheet
        create_summary_sheet(workbook, results_df, available_indices, unavailable_indices)
    
    print(f"✅ Excel file created: {filename}")
    print("\nFile contains:")
    print(f"  - Main sheet: CAGR Analysis ({len(results_df)} rows × {len(results_df.columns)} columns)")
    print(f"  - Summary sheet: Key statistics and insights")
    print("\nCAGR Analysis completed successfully!")
    
    return filename, results_df, available_indices, unavailable_indices

def apply_excel_formatting(worksheet, df):
    """Apply professional formatting to the Excel worksheet"""
    
    # Define colors
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    index_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    na_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Define fonts
    header_font = Font(bold=True, color="FFFFFF", size=12)
    index_font = Font(bold=True, size=10)
    data_font = Font(size=10)
    
    # Define alignment
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format headers (first row)
    for col in range(1, len(df.columns) + 2):  # +2 for index column
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # Format index column (first column) 
    for row in range(2, len(df) + 2):
        cell = worksheet.cell(row=row, column=1)
        cell.fill = index_fill
        cell.font = index_font
        cell.alignment = left_align
        cell.border = thin_border
    
    # Format data cells
    for row in range(2, len(df) + 2):
        for col in range(2, len(df.columns) + 2):
            cell = worksheet.cell(row=row, column=col)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border
            
            # Color code based on value
            if cell.value == "NA":
                cell.fill = na_fill
            elif isinstance(cell.value, (int, float)):
                if cell.value > 0:
                    cell.fill = positive_fill
                else:
                    cell.fill = negative_fill
                
                # Format as percentage
                cell.number_format = '0.00"%"'
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        if column_letter == 'A':  # Index names column
            adjusted_width = min(max_length + 2, 35)  # Wider for index names
        else:
            adjusted_width = min(max_length + 2, 15)  # Standard width for data
        worksheet.column_dimensions[column_letter].width = adjusted_width

def create_summary_sheet(workbook, df, available_indices, unavailable_indices):
    """Create a summary sheet with key statistics"""
    
    # Create new worksheet
    summary_sheet = workbook.create_sheet(title="Summary & Insights")
    
    # Calculate summary statistics
    summary_data = []
    
    # Header
    summary_data.append(["CAGR ANALYSIS SUMMARY", ""])
    summary_data.append(["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary_data.append(["", ""])
    
    # Basic statistics
    summary_data.append(["BASIC STATISTICS", ""])
    summary_data.append(["Total Indices Analyzed:", len(df)])
    summary_data.append(["Available Indices:", len(available_indices)])
    summary_data.append(["Unavailable Indices:", len(unavailable_indices)])
    summary_data.append(["Data Availability:", f"{len(available_indices)/len(df)*100:.1f}%"])
    summary_data.append(["Time Periods:", f"1 to 20 years"])
    summary_data.append(["", ""])
    
    # Best and worst performers for different periods
    periods_to_analyze = ["1Y", "3Y", "5Y", "10Y", "15Y", "20Y"]
    
    for period in periods_to_analyze:
        if period in df.columns:
            summary_data.append([f"TOP PERFORMERS - {period}", ""])
            
            # Get data for this period, excluding NA values
            period_data = df[period].replace("NA", np.nan).dropna()
            
            if not period_data.empty:
                # Top 5 performers
                top_performers = period_data.nlargest(5)
                for idx, (index_name, cagr) in enumerate(top_performers.items(), 1):
                    summary_data.append([f"  {idx}. {index_name}", f"{cagr:.2f}%"])
                
                summary_data.append(["", ""])
                
                # Bottom 5 performers
                summary_data.append([f"BOTTOM PERFORMERS - {period}", ""])
                bottom_performers = period_data.nsmallest(5)
                for idx, (index_name, cagr) in enumerate(bottom_performers.items(), 1):
                    summary_data.append([f"  {idx}. {index_name}", f"{cagr:.2f}%"])
            else:
                summary_data.append(["  No data available", ""])
            
            summary_data.append(["", ""])
    
    # Average CAGR by time period
    summary_data.append(["AVERAGE CAGR BY TIME PERIOD", ""])
    for period in df.columns:
        period_data = df[period].replace("NA", np.nan)
        avg_cagr = period_data.mean()
        if not np.isnan(avg_cagr):
            summary_data.append([f"  {period}:", f"{avg_cagr:.2f}%"])
        else:
            summary_data.append([f"  {period}:", "No data"])
    
    summary_data.append(["", ""])
    
    # List unavailable indices
    if unavailable_indices:
        summary_data.append(["UNAVAILABLE INDICES", ""])
        for name, symbol in unavailable_indices.items():
            summary_data.append([f"  {name}", symbol])
    
    # Write summary data to sheet
    for row_num, (col1, col2) in enumerate(summary_data, 1):
        summary_sheet.cell(row=row_num, column=1, value=col1)
        summary_sheet.cell(row=row_num, column=2, value=col2)
    
    # Format summary sheet
    summary_sheet.column_dimensions['A'].width = 40
    summary_sheet.column_dimensions['B'].width = 20
    
    # Apply formatting to headers
    for row in summary_sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if cell.value.isupper() and cell.column == 1:
                    cell.font = Font(bold=True, size=12)
                elif cell.value.startswith("  ") and ":" in cell.value:
                    cell.font = Font(size=10)
                elif ":" in cell.value and cell.column == 1:
                    cell.font = Font(bold=True, size=10)

if __name__ == "__main__":
    filename, results_df, available_indices, unavailable_indices = main()
    
    # Display a preview of the results
    print("\n" + "="*80)
    print("PREVIEW OF RESULTS")
    print("="*80)
    print(results_df.head(10))
    
    if len(results_df) > 10:
        print(f"\n... and {len(results_df) - 10} more rows")
    
    print(f"\nFull results saved to: {filename}") 